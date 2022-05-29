import os
import re
import yaml
from typing import Union
import openpyxl

class MyExcelLib:
    def __init__(self, file_path, file_name, path_to_config='./style/', style_file_name='style.yml'):
        """エクセルファイルを操作するクラス"""
        self._file_path = file_path if file_path[-1]=='/' else file_path+'/'
        self._file_name = file_name
        self._path_to_config = path_to_config if path_to_config[-1]=='/' else path_to_config+'/'
        self._style_file_name = style_file_name
        self._book = None

        assert os.path.exists(path_to_config + style_file_name), f'\'{style_file_name}\' file does not exist.'
        assert os.path.exists(file_path), f'\'{file_path}\' directory does not exist.'

        if os.path.exists(file_path + file_name):
            print(f'\'{self._file_path+self._file_name}\' file already exists.')
            print('Please run \'MyExcelLib.load_book()\' method.')
        else:
            print(f'\'{self._file_path+self._file_name}\' file does not exist.')
            print('Please run \'MyExcelLib.create_book()\' method.')

    # ブック
    def load_book(self):
        """ブックの読込み.
        インスタンスがすでにブックの情報を持っている場合、エラー"""
        assert self._book is None, '\'self._book\' is not None. You can\'t load a new book.'
        self._book = openpyxl.load_workbook(self._file_path + self._file_name)
        self.activate_sheet(self._book.sheetnames[0])

    def create_book(self):
        """新規ブックの作成.
        インスタンスがすでにブックの情報を持っている場合、エラー
        作成しようとしているファイルがすでに存在する場合、エラー
        """
        assert self._book is None, '\'self._book\' is not None. You can\'t create a new book.'
        assert os.path.exists(self._file_path + self._file_name), f'\'{self._file_name}\' file already exists. Please load or remove the file.'
        self._book = openpyxl.Workbook()
        self.activate_sheet(self._book.sheetnames[0])

    def save_book(self):
        """ブックの保存"""
        print(f'save {self._file_path+self._file_name}')
        self._book.save(self._file_path + self._file_name)
    
    # シート
    def set_sheet_name(self, name: str, index: int):
        """シート名の変更
        Args:
            name (str): 変更後のシート名.
            index (int): 変更対象のシートindex.
        """
        self._book.worksheets[index].title = name

    def add_sheet(self, title: str, index: int=None):
        """シートの追加
        Args:
            title (str): 追加するシート名.
            index (int, optional): 何番目に追加するか. Default to None.
        """
        self._book.create_sheet(title, index)

    def remove_sheet(self, title: str):
        """シートの削除
        Args:
            title (str): 削除するシート名.
        """
        self._book.remove(title)
    
    def activate_sheet(self, sheet):
        """アクティブシートの変更
        Args:
            sheet (str): アクティブにしたいシート名
        """
        assert sheet in self._book.sheetnames, f'\'{sheet}\' sheet does not exist.'
        self._active_sheet = sheet
        print(f'\'{sheet}\' sheet has been activated.')

    def _get_active_sheet(self):
        """アクティブシートの取得"""
        return self._book[self._active_sheet]

    # セル
    def set_value(self, value, position, style='basic'):
        """セルへ値の設定
        Args:
            value (str, int): セルへ入力したい値.
            sheet (str): 設定先のシート名.
            position (str, tuple): 設定先のセル. 'B3'の形式でも(3, 2)の形式でも可.
            style (str, optional): セルの書式. Default to basic.
        """
        sheet = self._get_active_sheet()
        position = get_cell_position(position)
        cell = sheet.cell(*position)
        cell.value = value
        self.set_style(position, style)

    def set_link2cell(self, value, to_sheet, to_position, position, style='hyperlink'):
        """同ブックセルへのリンク
        Args:
            value (str, int): リンクを設定する値.
            to_sheet (str): 移動先のシート名.
            to_position (str, tuple): 移動先のセル. 'B3'の形式も(3, 2)の形式でも可.
            sheet (str): 設定先のシート名.
            position (str, tuple): 設定先のセル. 'B3'の形式も(3, 2)の形式でも可.
            style (str, optional): セルの書式. Default to basic.
        """
        sheet = self._get_active_sheet()
        to_position = get_cell_position(to_position)
        position = get_cell_position(position)
        link = f'{self._file_name}\#{to_sheet}!{tuple2str(to_position)}'
        cell = sheet.cell(*position)
        cell.value = value
        cell.hyperlink = link
        self.set_style(position, style)

    def set_link2web(self, value, url, position, style='hyperlink'):
        """Webサイトへのリンク
        Args:
            value (str, int): リンクを設定する値.
            url (str): 移動先のURL.
            position (str, tuple): 設定先のセル. 'B3'の形式も(3, 2)の形式でも可.
            style (str, optional): セルの書式. Default to basic.
        """
        sheet = self._get_active_sheet()
        position = get_cell_position(position)
        cell = sheet.cell(*position)
        cell.value = value
        cell.hyperlink = url
        self.set_style(position, style)

    def get_value(self, position):
        """セルの値の取得
        Args:
            position (str, tuple): 設定先のセル. 'B3'の形式も(3, 2)の形式でも可.
        """
        sheet = self._get_active_sheet()
        position = get_cell_position(position)
        cell = sheet.cell(*position)
        return cell.value
    
    # 書式
    def set_style(self, position, style):
        """セルに対してスタイルの適用.
        結合セルに枠線を適用する場合、全てのセルに対して実行する必要がある.
        Args:
            position (str, tuple): 設定先のセル. 'B3'の形式も(3, 2)の形式でも可.
            style (str): セルの書式. Default to 'basic'.
        """
        sheet = self._get_active_sheet()
        position = get_cell_position(position)
        cell = sheet.cell(*position)
        style = self._import_style(style)

        if 'font' in style.keys():
            cell.font = openpyxl.styles.Font(**style['font'])
        if 'alignment' in style.keys():
            cell.alignment = openpyxl.styles.alignment.Alignment(**style['alignment'])
        if 'fill' in style.keys():
            cell.fill = openpyxl.styles.PatternFill(**style['fill'])
        if 'border' in style.keys():
            is_merged_cell = self._is_merged_cell(position)
            if is_merged_cell: # 結合セル
                target_cell = [sheet.cell(*p) for p in is_merged_cell]
            else: # 結合ではない
                target_cell = [cell]
            for cell in target_cell:
                border_style = {}
                for side in style['border'].keys():
                    border_style[side] = openpyxl.styles.Side(**style['border'][side])
                if 'diagonal' in border_style.keys():
                    border_style['diagonalDown'] = True
                cell.border = openpyxl.styles.Border(**border_style)
    
    def concat_cells(self, start_position, end_position):
        """セルの結合と値の設定、スタイルの適用.
        結合されたセルに別メソッドを用いてスタイルの適用をするとうまく適用されないため、このメソッドで行う.
        Args:
            start_position (str, tuple): 結合先頭のセル. 'B3'の形式も(3, 2)の形式でも可.
            end_position (str, tuple): 結合末尾のセル. 'B3'の形式も(3, 2)の形式でも可.
        """
        sheet = self._get_active_sheet()
        start_position = get_cell_position(start_position)
        end_position = get_cell_position(end_position)
        sheet.merge_cells(
            start_row=start_position[0], 
            start_column=start_position[1],
            end_row=end_position[0],
            end_column=end_position[1])

    def set_width(self, width, col):
        """列幅の指定
        Args:
            width (int|float): 列の幅
            col (str|int): 列. 'B'の形式もint形式でも可.
        """
        sheet = self._get_active_sheet()
        if type(col)==int:
            col = re.sub("\d", "", tuple2str((1, col)))
        sheet.column_dimensions[col].width = width

    def set_height(self, height, row):
        """行高の指定
        Args:
            height (int|float): 行の高さ
            row (int): 行.
        """
        sheet = self._get_active_sheet()
        sheet.row_dimensions[row].height = height

    def set_df(self, df, position):
        """dfの配置
        Args:
            df (pd.DataFrame): 設定したいdataframe.
            position (str, tuple): 設定先のセル. 'B3'の形式も(3, 2)の形式でも可.
        """
        sheet = self._get_active_sheet()
        position = get_cell_position(position)
        
        # df左上
        self.set_value('', position, 'df_head')
        for i, c in enumerate(df.columns):
            self.set_value(c, (position[0], position[1]+(i+1)), 'df_head')
        for i, idx in enumerate(df.index):
            self.set_value(idx, (position[0]+i+1, position[1]), 'df_value')
        for r in range(df.shape[0]):
            for c in range(df.shape[1]):
                self.set_value(df.iat[r,c], (position[0]+1+r, position[1]+1+c), 'df_value')

    def _is_merged_cell(self, position):
        """指定したセルが結合されたセルか判定を行う.
        Args:
            position (str, tuple): 判定を行うセル.
        Returns:
            (lis or bool): Yesなら結合しているセルの位置listを返す. NoならFalseを返す.
        """
        sheet = self._get_active_sheet()
        position = get_cell_position(position)
        ranges = sheet.merged_cells.ranges
        for merged_cell in ranges:
            start, end = str(merged_cell).split(':')
            start, end = str2tuple(start), str2tuple(end)

            condition1 = (start[0]>=position[0])&(end[0]>=position[0])
            condition2 = (start[1]>=position[1])&(end[1]>=position[1])
            if condition1 & condition2:
                merged_cells = []
                for row in range(start[0], end[0]+1):
                    for col in range(start[1], end[1]+1):
                        merged_cells.append((row, col))
                return merged_cells
        return False

    def _import_style(self, style)->dict:
        """設定ファイル(yml)を読み込んで返す.
        styleがbasicでない場合、basicに上書きして返す.
        Return:
            config(dict): 読み込んだ設定値
        """
        with open(self._path_to_config + self._style_file_name, 'r') as yml:
            styles = yaml.load(yml,  Loader=yaml.SafeLoader)

        style_dict = styles['basic']
        if style != 'basic':
            style_dict.update(styles[style]) # basicスタイルへの上書き
        return style_dict

def get_cell_position(position)->tuple:
    """セル位置を文字列で渡された場合に数値のタプルで返す.
    タプルを渡された場合も正常に動作するため、セル位置を扱う場合にはこの関数を通す.

    Args:
        position (str, tuple): 設定先のセル. 'B3'の形式も(3, 2)の形式でも可.
    Returns:
        tuple: タプル形式になったセル.
    """
    if type(position) == str:
        position = str2tuple(position)
    return (position[0], position[1])

def str2tuple(pos_str:str)->tuple:
    """セル'英字+数字'->(行番号, 列番号)
    列最小値が0でないため単純な26進数変換を適用できない"""
    al_num = lambda x: 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'.index(x) + 1
    
    row = int(re.sub("\D", "", pos_str))
    col_str = re.sub("\d", "", pos_str)
    col = sum([al_num(s)*(26**i) for i, s in enumerate(col_str[::-1])])
    return (row, col)

def tuple2str(pos_tuple:tuple)->str:
    """セル(行番号, 列番号)->'英字+数字'
    列最小値が0でないため単純な26進数変換を適用できない"""
    row, col = pos_tuple
    li = []
    while True:
        sub_flag = 0
        mod = col % 26
        if mod==0:
            sub_flag=1
            mod = 26
        li.append('ABCDEFGHIJKLMNOPQRSTUVWXYZ'[mod-1])
        col = col // 26 - sub_flag
        if col == 0:
            break
    col_str = ''.join(li[::-1])
    pos_str = col_str + str(row)
    return pos_str